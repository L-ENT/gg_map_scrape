import io
import threading
import time

import pandas as pd
from openpyxl import load_workbook

from app import (
    EXPORT_COLUMNS,
    AI_MODE_MIN_EVIDENCE_CHARS,
    AI_MODE_PENDING_MIN_EVIDENCE_CHARS,
    ClinicCache,
    Evidence,
    ai_mode_has_pending_marker,
    ai_mode_evidence_is_ready,
    apply_debug_keep_selection,
    append_rows_preserving_template,
    candidate_id,
    candidate_result,
    existing_lead_keys,
    early_maps_rejection,
    extract_operation_time_from_text,
    filter_result,
    lead_identity_keys,
    maps_place_id_from_url,
    merge_gemini_evidence,
    owner_role_is_explicit,
    promote_updater_payload,
    run_job,
    start_background_job,
    should_keep_in_final_output,
    wait_for_manual_captcha,
)


def test_extract_operation_time_from_relevant_sentence():
    page_text = "Monday-Friday 9:00 AM - 5:00 PM. We offer counseling services."
    assert extract_operation_time_from_text(page_text) == "Monday-Friday 9:00 AM - 5:00 PM"


def test_ai_mode_rejects_streaming_and_short_answers():
    partial = "Đang tìm kiếm " + ("clinic information " * 10)
    short_header = "Here is a brief, scannable overview of the clinic in Provo, Utah:"
    assert len(partial) < AI_MODE_PENDING_MIN_EVIDENCE_CHARS
    assert ai_mode_has_pending_marker(partial) is True
    assert ai_mode_evidence_is_ready(partial) is False
    assert ai_mode_evidence_is_ready(short_header) is False


def test_ai_mode_accepts_complete_answer_without_pending_marker():
    complete = "Owner Jane Doe. Private group counseling practice. " + ("Verified clinic details and services. " * 10)
    assert len(complete) >= AI_MODE_MIN_EVIDENCE_CHARS
    assert ai_mode_evidence_is_ready(complete) is True


def test_incomplete_ai_mode_is_not_sent_to_gemini(monkeypatch):
    row = {
        "Practice name": "Incomplete Clinic",
        "location": "Provo",
        "operation time and days": "N/A",
    }
    monkeypatch.setattr("app.maps_search_urls", lambda *args, **kwargs: [("https://maps.test/place", "Incomplete Clinic")])
    monkeypatch.setattr("app.extract_maps_place_with_retry", lambda *args, **kwargs: dict(row))
    ai_calls = []
    def incomplete_ai(*args, **kwargs):
        ai_calls.append(True)
        return ""
    monkeypatch.setattr("app.google_ai_overview", incomplete_ai)

    def unexpected_gemini_call(*args, **kwargs):
        raise AssertionError("Gemini must not be called without complete AI Mode evidence")

    monkeypatch.setattr("app.gemini_metadata", unexpected_gemini_call)
    known = set()
    accepted, debug = run_job(object(), "Provo", "UT", "therapy", 10, known, lambda message: None, "api-key")

    assert accepted == []
    assert debug[0]["Filter result"] == "RETRY: AI Mode chưa hoàn tất"
    assert debug[0]["Gemini"] == "not called"
    assert ("incomplete clinic", "provo") not in known
    assert len(ai_calls) == 2


def test_ai_mode_failure_is_retried_after_the_main_list(monkeypatch):
    rows = {
        "First Clinic": {"Practice name": "First Clinic", "location": "Provo", "operation time and days": "N/A"},
        "Second Clinic": {"Practice name": "Second Clinic", "location": "Provo", "operation time and days": "N/A"},
    }
    monkeypatch.setattr("app.maps_search_urls", lambda *args, **kwargs: [("first", "First Clinic"), ("second", "Second Clinic")])
    monkeypatch.setattr("app.extract_maps_place_with_retry", lambda driver, url, city, name, *args: dict(rows[name]))
    ai_order = []
    def ai_result(driver, name, *args, **kwargs):
        ai_order.append(name)
        if name == "First Clinic" and ai_order.count(name) == 1:
            return ""
        return "Complete AI Mode evidence " * 20
    monkeypatch.setattr("app.google_ai_overview", ai_result)
    monkeypatch.setattr("app.gemini_metadata", lambda *args, **kwargs: {
        "status": "ok", "owner": "N/A", "operation_time_and_days": "N/A",
        "doctor_count": 2, "branch_count": 1, "is_solo": False,
        "is_collective": False, "nonprofit": False, "private_practice": True,
        "target_service": True, "red_flags": [], "disallowed_provider_title": False,
        "outdated_or_insufficient": False, "over_25_years": False, "has_board": False,
    })
    candidate_progress = []
    retry_waiting_changes = []

    accepted, debug = run_job(
        object(), "Provo", "UT", "therapy", 10, set(), lambda message: None, "api-key",
        on_candidate_progress=lambda done, total: candidate_progress.append((done, total)),
        on_retry_waiting=retry_waiting_changes.append,
    )

    assert ai_order == ["First Clinic", "Second Clinic", "First Clinic"]
    assert {row["Practice name"] for row in accepted} == {"First Clinic", "Second Clinic"}
    assert all(item["Filter result"] != "RETRY: AI Mode chưa hoàn tất" for item in debug)
    assert candidate_progress == [(1, 2), (2, 2)]
    assert retry_waiting_changes == [1, -1]


def test_captcha_wait_can_be_stopped_without_a_timeout(monkeypatch):
    messages = []
    monkeypatch.setattr("app.captcha_is_visible", lambda driver: True)

    assert wait_for_manual_captcha(object(), messages.append, lambda: True) is False
    assert messages == ["Đã dừng trong khi chờ xác minh CAPTCHA."]


def test_debug_callback_receives_each_processed_clinic(monkeypatch):
    row = {"Practice name": "Live Clinic", "location": "Provo", "operation time and days": "N/A"}
    monkeypatch.setattr("app.maps_search_urls", lambda *args, **kwargs: [("https://maps.test/place", "Live Clinic")])
    monkeypatch.setattr("app.extract_maps_place_with_retry", lambda *args, **kwargs: dict(row))
    monkeypatch.setattr("app.google_ai_overview", lambda *args, **kwargs: "")
    updates = []

    run_job(object(), "Provo", "UT", "therapy", 10, set(), lambda message: None, "api-key", on_debug=updates.append)

    assert len(updates) == 1
    assert updates[0]["Practice"] == "Live Clinic"


def test_keep_selection_uses_visible_row_number_when_hidden_id_is_missing(monkeypatch):
    sheet = "Salt Lake City, UT"
    candidate = {"Practice name": "Capstone Counseling Centers", "location": "Salt Lake City"}
    item_id = candidate_id(sheet, candidate)
    saves = []
    monkeypatch.setattr("app.save_checkpoint_locked", lambda job: saves.append(True))
    job = {
        "lock": threading.Lock(),
        "debug": [{"Candidate ID": item_id, "Keep": False, "Sheet": sheet}],
        "candidates": {item_id: candidate},
        "rows_by_sheet": {sheet: []},
        "message": "",
    }
    edited = pd.DataFrame([{"STT": 1, "Keep": True, "Sheet": sheet}])

    assert apply_debug_keep_selection(job, edited) == 1
    assert job["rows_by_sheet"][sheet] == [candidate]
    assert job["debug"][0]["Keep"] is True
    assert saves == [True]


def test_keep_logic_and_debug_result_agree_for_matching_clinic():
    verification = {"Is_NonProfit_StateOwned": False, "Contains_Red_Flags": False, "Has_Multiple_Therapists": True, "Contains_Target_Services_Licenses": True, "Owner's name": "N/A"}
    assert should_keep_in_final_output(verification) is True
    assert filter_result(verification) == "KEEP"


def test_keep_logic_and_debug_result_agree_without_qualifying_evidence():
    verification = {"Is_NonProfit_StateOwned": False, "Contains_Red_Flags": False, "Has_Multiple_Therapists": False, "Contains_Target_Services_Licenses": False, "Owner's name": "N/A"}
    assert should_keep_in_final_output(verification) is False
    assert filter_result(verification) == "REJECT: insufficient qualifying evidence"


def test_solo_is_rejected_only_when_25_plus_years():
    base = {"is_solo_practitioner": True, "Contains_Target_Services_Licenses": True}
    assert filter_result({**base, "mentions_25_plus_years_experience": False}) == "KEEP"
    assert filter_result({**base, "mentions_25_plus_years_experience": True}) == "REJECT: solo practitioner with 25+ years experience"
    assert filter_result({"is_solo_practitioner": False, "mentions_25_plus_years_experience": True, "Contains_Target_Services_Licenses": True}) == "KEEP"


def test_collective_with_direct_therapist_phone_can_be_kept():
    collective = {"is_therapist_collective_or_independent": True, "Contains_Target_Services_Licenses": True}
    assert filter_result({**collective, "Has_Direct_Therapist_Phone": False}) == "REJECT: therapist collective without direct therapist phone"
    assert filter_result({**collective, "Has_Direct_Therapist_Phone": True}) == "KEEP"


def test_collective_phone_exception_does_not_bypass_other_rejection_rules():
    collective = {
        "is_therapist_collective_or_independent": True,
        "Has_Direct_Therapist_Phone": True,
        "Contains_Target_Services_Licenses": True,
        "Contains_Disallowed_Provider_Title": True,
    }
    assert filter_result(collective) == "REJECT: MD / DO / PMHNP"


def test_direct_therapist_phone_is_propagated_to_filter_evidence():
    evidence = merge_gemini_evidence(Evidence(), {"status": "ok", "direct_therapist_phone": True})
    assert evidence.direct_therapist_phone is True


def test_rejects_only_above_sixty_therapists_or_five_locations():
    qualifying = {"Contains_Target_Services_Licenses": True}
    assert filter_result({**qualifying, "doctor_count": 60}) == "KEEP"
    assert filter_result({**qualifying, "doctor_count": 61}) == "REJECT: >60 therapists"
    assert filter_result({**qualifying, "branch_count": 5}) == "KEEP"
    assert filter_result({**qualifying, "branch_count": 6}) == "REJECT: >5 locations"


def test_owner_role_must_explicitly_prove_ownership():
    assert owner_role_is_explicit("Co-Founder and CEO") is True
    assert owner_role_is_explicit("Clinical Director") is False
    assert owner_role_is_explicit("Likely owner") is False
    assert owner_role_is_explicit("Former founder") is False


def test_identity_keys_match_name_phone_and_official_domain():
    first = lead_identity_keys({"Practice Name": "Example Therapy, PLLC", "Phone Number": "+1 (801) 555-1212", "Website Link": "https://www.exampletherapy.com/team"})
    second = lead_identity_keys({"Practice name": "Example Therapy LLC", "phone number": "8015551212", "website link": "https://exampletherapy.com/contact"})
    assert first == second == {"name:example therapy", "phone:8015551212", "domain:exampletherapy.com"}


def test_shared_profile_subdomain_is_not_used_as_an_organization_identity():
    keys = lead_identity_keys({"Practice Name": "Example Therapy", "Website Link": "https://m.facebook.com/exampletherapy"})
    assert keys == {"name:example therapy"}


def test_uploaded_workbook_without_owner_builds_global_duplicate_keys():
    source = io.BytesIO()
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        pd.DataFrame([{"Practice Name": "Existing Clinic, LLC", "Phone Number": "801-555-9999", "Website Link": "https://existing.example"}]).to_excel(writer, sheet_name="Provo, UT", index=False)
    keys = existing_lead_keys(source.getvalue())
    assert {"name:existing clinic", "phone:8015559999", "domain:existing.example"} <= keys


def test_duplicate_phone_skips_ai_mode(monkeypatch):
    row = {"Practice name": "Renamed Clinic", "phone number": "801-555-1212", "website link": "N/A", "location": "Provo", "operation time and days": "N/A"}
    monkeypatch.setattr("app.maps_search_urls", lambda *args, **kwargs: [("place", "Different Maps Name")])
    monkeypatch.setattr("app.extract_maps_place_with_retry", lambda *args, **kwargs: dict(row))
    monkeypatch.setattr("app.google_ai_overview", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("Duplicate must skip AI Mode")))
    _, debug = run_job(object(), "Provo", "UT", "therapy", 10, {"phone:8015551212"}, lambda message: None, "api-key")
    assert debug[0]["Filter result"] == "SKIP: duplicate lead"


def test_place_id_and_address_are_part_of_duplicate_identity():
    keys = lead_identity_keys({"Practice Name": "Example Therapy", "maps place id": "ChIJ123", "address": "10 Main St., Provo, UT"})
    assert "place:ChIJ123" in keys
    assert "name-address:example therapy|10 main st provo ut" in keys
    assert maps_place_id_from_url("https://maps.google.com/maps/place/Test/data=!4m2!3m1!1sChIJ123!8m2") == "ChIJ123"


def test_maps_early_rejection_is_limited_to_explicit_status_or_category():
    assert early_maps_rejection({"maps raw text": "Permanently closed"}) == "permanently closed on Google Maps"
    assert early_maps_rejection({"maps category": "Psychiatric hospital"}) == "Google Maps category: psychiatric hospital"
    assert early_maps_rejection({"maps category": "Mental health clinic", "maps raw text": "We refer medical treatment elsewhere"}) == ""


def test_sqlite_cache_persists_all_evidence_layers(tmp_path):
    cache = ClinicCache(tmp_path / "clinic_cache.sqlite")
    row = {"Practice name": "Cached Clinic", "maps place id": "ChIJCACHE", "phone number": "8015551212", "website link": "https://cached.example", "location": "Provo", "maps raw text": "Maps evidence"}
    metadata = {"status": "ok", "owner": "Jane Doe", "owner_role": "Founder", "doctor_count": 3}
    cache.save(row)
    cache.save(row, ai_evidence="Complete AI Mode evidence", metadata=metadata)

    restored = ClinicCache(tmp_path / "clinic_cache.sqlite").lookup({"maps place id": "ChIJCACHE"})

    assert restored["maps"]["maps raw text"] == "Maps evidence"
    assert restored["ai_evidence"] == "Complete AI Mode evidence"
    assert restored["gemini"]["owner"] == "Jane Doe"


def test_cache_does_not_mix_same_name_in_different_cities(tmp_path):
    cache = ClinicCache(tmp_path / "clinic_cache.sqlite")
    cache.save({"Practice name": "Common Counseling", "location": "Provo", "maps raw text": "Provo evidence"}, ai_evidence="Provo AI")

    assert cache.lookup({"Practice name": "Common Counseling", "location": "Orem"}) is None


def test_cached_evidence_skips_maps_ai_and_gemini(monkeypatch, tmp_path):
    cache = ClinicCache(tmp_path / "clinic_cache.sqlite")
    row = {"Practice name": "Cached Clinic", "maps place id": "ChIJCACHE", "phone number": "8015551212", "website link": "https://cached.example", "location": "Provo", "operation time and days": "N/A", "maps raw text": "Maps evidence"}
    metadata = {"status": "ok", "owner": "Jane Doe", "owner_role": "Founder", "doctor_count": 1, "branch_count": 1, "target_service": True, "red_flags": []}
    cache.save(row, ai_evidence="Complete cached AI evidence", metadata=metadata)
    monkeypatch.setattr("app.maps_search_urls", lambda *args, **kwargs: [("https://maps.test/data=!1sChIJCACHE!8m2", "Cached Clinic")])
    monkeypatch.setattr("app.extract_maps_place_with_retry", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("Maps detail must use cache")))
    monkeypatch.setattr("app.google_ai_overview", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("AI Mode must use cache")))
    monkeypatch.setattr("app.gemini_metadata", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("Gemini must use cache")))

    accepted, debug = run_job(object(), "Provo", "UT", "therapy", 10, set(), lambda message: None, "api-key", cache=cache)

    assert accepted[0]["Practice name"] == "Cached Clinic"
    assert debug[0]["Gemini"] == "ok"


def test_rate_limit_saves_ai_evidence_instead_of_rejecting(monkeypatch, tmp_path):
    cache = ClinicCache(tmp_path / "clinic_cache.sqlite")
    row = {"Practice name": "Quota Clinic", "maps place id": "ChIJQUOTA", "phone number": "8015550000", "website link": "N/A", "location": "Provo", "operation time and days": "N/A", "maps raw text": "Maps evidence"}
    monkeypatch.setattr("app.maps_search_urls", lambda *args, **kwargs: [("place", "Quota Clinic")])
    monkeypatch.setattr("app.extract_maps_place_with_retry", lambda *args, **kwargs: dict(row))
    monkeypatch.setattr("app.google_ai_overview", lambda *args, **kwargs: "Complete AI Mode evidence")
    monkeypatch.setattr("app.gemini_metadata", lambda *args, **kwargs: {"status": "rate_limited"})

    _, debug = run_job(object(), "Provo", "UT", "therapy", 10, set(), lambda message: None, "api-key", cache=cache)

    assert debug[0]["Filter result"] == "WAITING: Gemini quota"
    assert cache.lookup({"maps place id": "ChIJQUOTA"})["ai_evidence"] == "Complete AI Mode evidence"


def test_two_browser_collectors_feed_one_gemini_thread(monkeypatch, tmp_path):
    browser_threads, result_threads = [], []
    barrier = threading.Barrier(2)

    class FakeDriver:
        def quit(self): pass

    def fake_build_driver(headless):
        browser_threads.append(threading.current_thread().name)
        return FakeDriver()

    def fake_run_job(driver, city, state, keyword, limit, known, progress, gemini_api_key, **kwargs):
        barrier.wait(timeout=2)
        row = {"Practice name": f"{city} Clinic", "location": city, "operation time and days": "N/A"}
        metadata = {"status": "ok", "owner": "N/A", "doctor_count": 1, "target_service": True, "red_flags": []}
        kwargs["on_ai_ready"](row, "Complete evidence", metadata)
        return [], []

    def tracked_result(*args, **kwargs):
        result_threads.append(threading.current_thread().name)
        return candidate_result(*args, **kwargs)

    monkeypatch.setattr("app.checkpoint_directory", lambda: tmp_path)
    monkeypatch.setattr("app.build_driver", fake_build_driver)
    monkeypatch.setattr("app.run_job", fake_run_job)
    monkeypatch.setattr("app.candidate_result", tracked_result)
    job = start_background_job({
        "source": b"", "jobs": [("Provo", "UT", "Provo, UT"), ("Orem", "UT", "Orem, UT")],
        "keywords": ["therapy"], "limit": 10, "parallel_workers": 2, "headless": False,
        "known": set(), "gemini_api_key": "key", "cache_path": tmp_path / "cache.sqlite",
    })
    deadline = time.time() + 5
    while job["running"] and time.time() < deadline:
        time.sleep(0.01)

    assert job["running"] is False
    assert set(browser_threads) == {"clinic-browser-1", "clinic-browser-2"}
    assert result_threads == ["clinic-gemini", "clinic-gemini"]


def test_append_rows_preserves_existing_data_and_adds_missing_columns():
    source = io.BytesIO()
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        pd.DataFrame([{"practice name": "Existing Clinic", "Phone Number": "111"}]).to_excel(writer, sheet_name="Provo, UT", index=False)
    new_rows = pd.DataFrame(
        [["New Clinic", "https://example.com", "222", "Provo", "9:00 AM - 5:00 PM", "4.8", "Jane Doe"]],
        columns=EXPORT_COLUMNS,
    )

    result = append_rows_preserving_template(source.getvalue(), {"Provo, UT": new_rows})
    workbook = load_workbook(io.BytesIO(result))
    sheet = workbook["Provo, UT"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]

    assert sheet.cell(2, 1).value == "Existing Clinic"
    assert sheet.cell(2, 2).value == "111"
    assert all(any(str(value).casefold() == header.casefold() for value in headers) for header in EXPORT_COLUMNS)
    assert sheet.max_row == 3


def test_updater_payload_replaces_installed_updater(tmp_path, monkeypatch):
    updater = tmp_path / "Clinic Lead Updater.exe"
    payload = tmp_path / "Clinic Lead Updater Payload.exe"
    updater.write_bytes(b"old updater")
    payload.write_bytes(b"new updater")
    monkeypatch.setattr("app.installed_app_directory", lambda: tmp_path)
    monkeypatch.setattr("app.sys.frozen", True, raising=False)
    monkeypatch.setattr("app.sys.platform", "win32")

    promote_updater_payload()

    assert updater.read_bytes() == b"new updater"
    assert not payload.exists()
