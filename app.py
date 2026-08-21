"""Google Maps + Google AI Overview clinic lead collector.

The collector never downloads a clinic website. Google Maps supplies contact
facts; the visible Google Search AI Overview supplies eligibility evidence.
"""
import io
import json
import os
import re
import subprocess
import sys
import time
import threading
import uuid
from queue import Empty, Queue
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import quote_plus, unquote_plus, urlparse

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

try:
    import winsound
except ImportError:
    winsound = None

st.set_page_config(page_title="Google Maps AI Overview clinic leads", page_icon="🗺️", layout="wide")
MAX_DOCTOR_COUNT, MAX_BRANCH_COUNT = 35, 5
# Always give the Maps result feed a meaningful chance to load.  After the
# minimum number of scrolls, stop only if it has remained unchanged for this
# many more passes (or Google explicitly says the list has ended).
MIN_MAPS_SCROLL_ROUNDS, MAPS_STALL_ROUNDS = 20, 20
MAX_SAVED_CHECKPOINTS = 10
GEMINI_MIN_REQUEST_INTERVAL_SECONDS = 5.0
GEMINI_MAX_RETRIES = 4
AI_MODE_MAX_ATTEMPTS = 2
AI_MODE_WAIT_SECONDS = 30
AI_MODE_STABLE_SECONDS = 2.0
AI_MODE_PENDING_STABLE_SECONDS = 5.0
AI_MODE_MIN_EVIDENCE_CHARS = 200
AI_MODE_PENDING_MIN_EVIDENCE_CHARS = 500
EXPORT_COLUMNS = ["Practice Name", "Website Link", "Phone Number", "Location (City Ne Only)", "Operation Time and Days", "Rating Star", "Owner"]
RED_FLAG_TERMS = ("intensive outpatient", "substance abuse", "addiction treatment", "medical treatment", "peer support", "medication management", "case management", "psychiatric hospital")
GEMINI_MODEL = "gemini-3.5-flash-lite"
UPDATE_REPOSITORY = "L-ENT/gg_map_scrape"
UPDATE_ASSET_NAMES = {
    "win32": "Clinic-Lead-Collector-windows.zip",
    "darwin": "Clinic-Lead-Collector-macos-arm64.zip",
}
_GEMINI_CACHE: Dict[str, Dict[str, Any]] = {}
_GEMINI_CACHE_LOCK = threading.Lock()
_GEMINI_RATE_LOCK = threading.Lock()
_GEMINI_NEXT_REQUEST_AT = 0.0

@dataclass
class Evidence:
    owner: str = "N/A"; doctor_count: Optional[int] = None; branch_count: Optional[int] = None
    is_solo: bool = False; is_collective: bool = False; old_or_insufficient: bool = False
    over_25_years: bool = False; has_board: bool = False; nonprofit: bool = False
    private_practice: Optional[bool] = None; disallowed_provider_title: bool = False
    red_flags: List[str] = field(default_factory=list); target_service: bool = False

def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()

def normalize_website(value: str) -> str:
    value = normalize_text(value)
    if not value or value.upper() in {"N/A", "LINK", "WEBSITE"}: return "N/A"
    if value.startswith("www."): value = "https://" + value
    parsed = urlparse(value)
    return value if parsed.scheme in {"http", "https"} and parsed.netloc else "N/A"

def extract_operation_time_from_text(text: str) -> str:
    compact = normalize_text(text)
    if re.search(r"\b(?:open )?24 hours\b", compact, re.I): return "24 hours"
    # AM/PM is mandatory: without it a phone fragment such as 85-23 is mistaken for hours.
    pat = r"(?:(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-z]*(?:\s*[-,&/]\s*(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-z]*)*\s*)?(?:0?[1-9]|1[0-2])(?::[0-5]\d)?\s*(?:AM|PM)\s*(?:-|–|—|to)\s*(?:0?[1-9]|1[0-2])(?::[0-5]\d)?\s*(?:AM|PM)"
    match = re.search(pat, compact, re.I)
    return match.group(0).strip() if match else "N/A"

def clean_owner_name(value: str) -> str:
    value = re.sub(r"^(?:Dr\.?|Mr\.?|Ms\.?|Mrs\.?)\s+", "", normalize_text(value), flags=re.I)
    value = re.sub(r",\s*(?:LCSW|LMFT|LPC|LMHC|PsyD|PhD|MSW|MA|MS)\b.*$", "", value, flags=re.I).strip(" ,.-")
    parts = value.replace(".", "").split()
    role_words = {"founder", "owner", "ceo", "director", "clinical", "chief", "executive", "unknown", "n/a", "ai", "mode", "overview", "google"}
    if len(parts) < 2 or any(part.lower() in role_words for part in parts): return "N/A"
    if not all(re.fullmatch(r"[A-Za-zÀ-ÿ'’-]+", part) for part in parts): return "N/A"
    return value if all(part[0].isupper() for part in parts) else "N/A"

def should_keep_in_final_output(v: Dict[str, Any]) -> bool:
    """Use the same decision text shown in Debug for the exported result."""
    return filter_result(v) == "KEEP"

def filter_result(v: Dict[str, Any]) -> str:
    """Human-readable reason shown in Debug for every accepted/rejected lead."""
    reasons = []
    checks = (("Is_NonProfit_StateOwned", "nonprofit / government"), ("Contains_Red_Flags", "disallowed service"), ("Contains_Disallowed_Provider_Title", "MD / DO / PMHNP"), ("is_private_practice", "not a private practice"), ("is_solo_practitioner", "solo practitioner"), ("seems_outdated_or_insufficient_website", "outdated/insufficient evidence"), ("mentions_25_plus_years_experience", "25+ years experience"), ("is_therapist_collective_or_independent", "therapist collective"), ("has_board_of_directors", "board of directors"))
    for key, label in checks:
        if key == "is_private_practice":
            if v.get(key) is False: reasons.append(label)
        elif v.get(key) is True: reasons.append(label)
    if isinstance(v.get("doctor_count"), (int, float)) and v["doctor_count"] >= MAX_DOCTOR_COUNT: reasons.append(f"{MAX_DOCTOR_COUNT}+ therapists")
    if isinstance(v.get("branch_count"), (int, float)) and v["branch_count"] > MAX_BRANCH_COUNT: reasons.append(f">{MAX_BRANCH_COUNT} locations")
    if reasons:
        return "REJECT: " + "; ".join(reasons)
    has_qualifying_evidence = bool(
        v.get("Contains_Target_Services_Licenses")
        or v.get("Has_Multiple_Therapists")
        or v.get("Owner's name", "N/A") != "N/A"
    )
    return "KEEP" if has_qualifying_evidence else "REJECT: insufficient qualifying evidence"

def evidence_as_verification(e: Evidence) -> Dict[str, Any]:
    return {"Owner's name": e.owner, "doctor_count": e.doctor_count, "branch_count": e.branch_count, "Is_NonProfit_StateOwned": e.nonprofit, "is_private_practice": e.private_practice, "Contains_Red_Flags": bool(e.red_flags), "Contains_Disallowed_Provider_Title": e.disallowed_provider_title, "is_solo_practitioner": e.is_solo, "seems_outdated_or_insufficient_website": e.old_or_insufficient, "mentions_25_plus_years_experience": e.over_25_years, "is_therapist_collective_or_independent": e.is_collective, "has_board_of_directors": e.has_board, "Contains_Target_Services_Licenses": e.target_service, "Has_Multiple_Therapists": bool(e.doctor_count and e.doctor_count > 1)}

def build_driver(headless: bool) -> webdriver.Chrome:
    options = Options()
    if headless: options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1200"); options.add_argument("--disable-notifications"); options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"]); options.add_experimental_option("useAutomationExtension", False)
    try: return webdriver.Chrome(options=options)
    except WebDriverException: return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

def body_text(driver: webdriver.Chrome) -> str:
    try: return normalize_text(driver.find_element(By.TAG_NAME, "body").text)
    except WebDriverException: return ""

def maybe_accept_google_consent(driver: webdriver.Chrome) -> None:
    for label in ("Accept all", "I agree", "Chấp nhận tất cả", "Đồng ý"):
        try: driver.find_element(By.XPATH, f"//*[self::button or @role='button'][contains(., '{label}')]").click(); time.sleep(.5); return
        except WebDriverException: pass

def captcha_is_visible(driver: webdriver.Chrome) -> bool:
    """Detect an actual Google verification challenge; this never attempts to solve it."""
    try:
        challenge_frames = driver.find_elements(By.CSS_SELECTOR, "iframe[src*='recaptcha'], iframe[title*='reCAPTCHA']")
        if challenge_frames: return True
        text = body_text(driver).lower()
        markers = ("i'm not a robot", "unusual traffic", "verify you are human", "xác minh bạn là người", "không phải là rô-bốt")
        return any(marker in text for marker in markers)
    except WebDriverException:
        return False

def wait_for_manual_captcha(driver: webdriver.Chrome, status: Any, timeout_seconds: int) -> bool:
    """Pause automation so the person at the visible Chrome window can verify."""
    if not captcha_is_visible(driver): return True
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        remaining = max(0, int(deadline - time.time()))
        status(f"Google đang yêu cầu CAPTCHA — hãy xác minh thủ công trong cửa sổ Chrome ({remaining}s còn lại).")
        time.sleep(2)
        if not captcha_is_visible(driver):
            status("Đã xác minh CAPTCHA, tiếp tục quét…")
            return True
    status("Hết thời gian chờ CAPTCHA; bỏ qua trang này và tiếp tục các lead khác.")
    return False

def maps_results_exhausted(driver: webdriver.Chrome) -> bool:
    """Return True only for Google Maps' explicit no-results/end-of-list UI."""
    try:
        text = normalize_text(driver.find_element(By.TAG_NAME, "body").text).lower()
    except WebDriverException:
        return False
    signals = (
        "you've reached the end of the list", "you’ve reached the end of the list",
        "reached the end of the list", "no results found", "no results",
        "đã đến cuối danh sách", "không tìm thấy kết quả", "không có kết quả",
    )
    return any(signal in text for signal in signals)

def maps_card_name_hint(card: Any, href: str) -> str:
    """Get the clinic name while it is still visible in the Maps results feed."""
    try:
        label = normalize_text(card.get_attribute("aria-label"))
        if label: return label
        lines = [normalize_text(line) for line in (card.text or "").splitlines() if normalize_text(line)]
        if lines: return lines[0]
    except WebDriverException:
        pass
    match = re.search(r"/maps/place/([^/@?]+)", href)
    return normalize_text(unquote_plus(match.group(1))) if match else "N/A"

def maps_search_urls(driver: webdriver.Chrome, city: str, state: str, keyword: str, limit: int, status: Any, captcha_wait_seconds: int) -> List[Tuple[str, str]]:
    query = f"{keyword} in {city}, {state}, USA" if state else f"{keyword} in {city}"
    driver.get(f"https://www.google.com/maps/search/{quote_plus(query)}"); time.sleep(2); maybe_accept_google_consent(driver)
    if not wait_for_manual_captcha(driver, status, captcha_wait_seconds): return []
    places: List[Tuple[str, str]] = []; seen_urls = set(); stale = 0; scroll_rounds = 0
    while len(places) < limit:
        before = len(places)
        for card in driver.find_elements(By.CSS_SELECTOR, "a[href*='/maps/place/']"):
            href = card.get_attribute("href") or ""
            if href and href not in seen_urls:
                places.append((href, maps_card_name_hint(card, href))); seen_urls.add(href)
            if len(places) >= limit: break
        stale = stale + 1 if len(places) == before else 0
        if maps_results_exhausted(driver):
            status(f"{city}: Google Maps báo đã hết kết quả ({len(places)} địa điểm).")
            break
        if scroll_rounds >= MIN_MAPS_SCROLL_ROUNDS and stale >= MAPS_STALL_ROUNDS:
            status(f"{city}: không có địa điểm mới sau {scroll_rounds} lượt cuộn; dừng tìm Maps.")
            break
        try:
            # Google Maps keeps results in a nested scrollable feed; scrolling
            # window/body alone leaves the search at its first handful of cards.
            feeds = driver.find_elements(By.CSS_SELECTOR, "div[role='feed']")
            target = feeds[0] if feeds else driver.find_element(By.TAG_NAME, "body")
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", target)
        except WebDriverException: break
        scroll_rounds += 1
        time.sleep(.6)
    return places[:limit]

def maps_value(driver: webdriver.Chrome, selectors: Sequence[str]) -> str:
    for selector in selectors:
        try:
            element = driver.find_element(By.CSS_SELECTOR, selector); value = normalize_text(element.get_attribute("aria-label") or element.text)
            if value: return value
        except WebDriverException: continue
    return "N/A"

def extract_maps_place(driver: webdriver.Chrome, url: str, city: str, name_hint: str = "N/A") -> Dict[str, Any]:
    driver.get(url); WebDriverWait(driver, 10).until(lambda d: d.execute_script("return document.readyState") in ("interactive", "complete")); time.sleep(.8)
    # Maps is a client-rendered app: the title is often delayed, but some valid
    # place views do not expose it through h1.  This is deliberately a soft
    # wait; lack of h1 must not discard the entire listing.
    try:
        WebDriverWait(driver, 6).until(lambda d: normalize_text(d.find_element(By.CSS_SELECTOR, "h1.DUwDvf, h1").text) not in {"", "Google Maps"})
    except TimeoutException:
        pass
    raw = body_text(driver); name = maps_value(driver, ("h1.DUwDvf", "h1")); name = name if name != "N/A" else name_hint; phone = maps_value(driver, ("button[data-item-id^='phone:']", "[aria-label^='Phone:']"))
    phone = re.sub(r"^(?:Phone|Điện thoại):\s*", "", phone, flags=re.I); website = "N/A"
    try: website = normalize_website(driver.find_element(By.CSS_SELECTOR, "a[data-item-id='authority']").get_attribute("href"))
    except WebDriverException: pass
    rating = maps_value(driver, ("div.F7nice span[aria-label*='star']", "span[aria-label*='star']")); match = re.search(r"\b(\d(?:\.\d)?)\s*(?:stars?|sao)\b", rating, re.I)
    return {"Practice name": name, "phone number": phone, "website link": website, "location": city, "operation time and days": extract_operation_time_from_text(raw), "rating star": match.group(1) if match else "N/A", "maps raw text": raw}

def extract_maps_place_with_retry(driver: webdriver.Chrome, url: str, city: str, name_hint: str, status: Any, captcha_wait_seconds: int) -> Dict[str, Any]:
    """Retry one transient Maps page-load failure; CAPTCHA remains manual."""
    last_error: Optional[Exception] = None
    for attempt in range(2):
        try:
            place = extract_maps_place(driver, url, city, name_hint)
            if place["Practice name"] != "N/A": return place
            if captcha_is_visible(driver):
                wait_for_manual_captcha(driver, status, captcha_wait_seconds)
            if attempt == 0:
                status(f"{city}: Maps chưa hiện tên địa điểm, tải lại một lần…")
                continue
            return place
        except (TimeoutException, WebDriverException) as exc:
            last_error = exc
            if captcha_is_visible(driver):
                wait_for_manual_captcha(driver, status, captcha_wait_seconds)
            if attempt == 0:
                status(f"{city}: lỗi tải Maps tạm thời, thử lại địa điểm này một lần…")
                time.sleep(1)
    raise last_error or WebDriverException("Could not load Google Maps place")

def _google_ai_overview_once(driver: webdriver.Chrome, name: str, city: str, state: str, status: Any, captcha_wait_seconds: int) -> str:
    """Use one focused AI Mode request to gather every lead-screening fact."""
    query = f'"{name}" {city} {state} briefly list owner, hours, private/nonprofit/government status, therapist count, locations, services including IOP, addiction, medication management, case management, peer support, medical treatment, solo/collective, board, 25+ years, MD/DO/PMHNP'
    driver.get(f"https://www.google.com/search?udm=50&q={quote_plus(query)}"); time.sleep(2.2); maybe_accept_google_consent(driver)
    if not wait_for_manual_captcha(driver, status, captcha_wait_seconds): return ""
    # `udm=50` is Google Search's AI Mode surface. Streaming answers can exceed
    # the old 15-second timeout, and a partially rendered answer can already be
    # longer than a sentence. Require the cleaned answer to stop changing for a
    # few seconds; when streaming chrome remains visible, use a stricter minimum
    # length and a longer stability window.
    try:
        main = WebDriverWait(driver, 20).until(lambda d: d.find_element(By.CSS_SELECTOR, "div[role='main']"))
    except (TimeoutException, WebDriverException):
        return ""
    status(f"{name}: AI Mode đang tạo câu trả lời đầy đủ…")
    deadline = time.monotonic() + AI_MODE_WAIT_SECONDS
    last_evidence = ""
    unchanged_since = time.monotonic()
    last_expand_at = 0.0
    while time.monotonic() < deadline:
        try:
            now = time.monotonic()
            if now - last_expand_at >= 2.0:
                expand_ai_mode_answer(driver, main)
                last_expand_at = now
            main_text = normalize_text(main.text)
            evidence = clean_ai_mode_evidence(main_text, query)
            if evidence != last_evidence:
                last_evidence = evidence
                unchanged_since = now
            else:
                required_stability = AI_MODE_PENDING_STABLE_SECONDS if ai_mode_has_pending_marker(evidence) else AI_MODE_STABLE_SECONDS
                if ai_mode_evidence_is_ready(evidence) and now - unchanged_since >= required_stability:
                    return evidence[:12000]
        except (WebDriverException, StaleElementReferenceException):
            try:
                main = driver.find_element(By.CSS_SELECTOR, "div[role='main']")
            except WebDriverException:
                return ""
        time.sleep(0.5)
    return ""

def google_ai_overview(driver: webdriver.Chrome, name: str, city: str, state: str, status: Any, captcha_wait_seconds: int) -> str:
    """Retry once when AI Mode is loading or explicitly has no answer."""
    for attempt in range(AI_MODE_MAX_ATTEMPTS):
        evidence = _google_ai_overview_once(driver, name, city, state, status, captcha_wait_seconds)
        if evidence: return evidence
        if attempt < AI_MODE_MAX_ATTEMPTS - 1:
            status(f"{name}: AI Mode chưa có câu trả lời hợp lệ, đang chạy lại clinic này một lần…")
    return ""

def expand_ai_mode_answer(driver: webdriver.Chrome, main: Any) -> None:
    """Open the visible AI Mode answer's disclosure controls before reading it."""
    button_xpath = ".//*[self::button or @role='button'][contains(normalize-space(.), 'Hiện tất cả') or contains(normalize-space(.), 'Show all') or contains(normalize-space(.), 'Xem thêm')]"
    for _ in range(3):
        clicked = False
        for button in main.find_elements(By.XPATH, button_xpath):
            try:
                if button.is_displayed() and button.is_enabled():
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();", button)
                    clicked = True
            except WebDriverException:
                continue
        if not clicked: break
        time.sleep(.35)

def clean_ai_mode_evidence(text: str, query: str) -> str:
    """Remove Google AI Mode chrome and echoed prompt before applying lead rules.

    The words in our question (for example `nonprofit` and `government`) are
    not facts about the clinic and must never trigger a rejection.
    """
    cleaned = normalize_text(text).replace("“", '"').replace("”", '"').replace("’", "'")
    # Google may echo our question with slightly different spacing/punctuation.
    query_pattern = re.escape(query).replace(r"\ ", r"\s+").replace(r"\,", r"\s*,?\s*")
    cleaned = re.sub(query_pattern, "", cleaned, flags=re.I)
    cleaned = re.sub(r"^(?:Chế độ AI|AI Mode)\s+(?:Tất cả|All)(?:\s+Hình ảnh|\s+Images)?.*?(?:Cuộc trò chuyện ở Chế độ AI:|AI Mode conversation:)", "", cleaned, flags=re.I)
    cleaned = re.sub(r"^(?:Cuộc trò chuyện ở Chế độ AI|AI Mode conversation)\s*:\s*", "", cleaned, flags=re.I)
    # The button's label can remain in Selenium's container text even after it
    # has been clicked. It is UI chrome, never clinic evidence.
    cleaned = re.sub(r"\b(?:Hiện tất cả|Show all|Xem thêm)\b", "", cleaned, flags=re.I)
    return normalize_text(cleaned)

def ai_mode_evidence_is_ready(evidence: str) -> bool:
    """Reject temporary and explicit-no-answer AI Mode screens as non-evidence."""
    compact = normalize_text(evidence).lower().strip(" .…")
    no_answer = (
        "không có câu trả lời nào cho nội dung tìm kiếm này",
        "hãy thử hỏi câu khác",
        "no answer for this search",
        "there are no answers for this search",
        "try asking something else",
    )
    minimum_length = AI_MODE_PENDING_MIN_EVIDENCE_CHARS if ai_mode_has_pending_marker(compact) else AI_MODE_MIN_EVIDENCE_CHARS
    return (
        len(compact) >= minimum_length
        and not any(marker in compact for marker in no_answer)
    )

def ai_mode_has_pending_marker(evidence: str) -> bool:
    """Detect Google streaming chrome, which may remain visible after completion."""
    prefix = normalize_text(evidence).lower()[:200]
    pending = ("đang tìm kiếm", "searching", "đang suy nghĩ", "thinking", "generating", "đang tạo", "loading")
    return any(marker in prefix for marker in pending)

def wait_for_gemini_slot() -> None:
    """Share one conservative Gemini request cadence across all Chrome workers."""
    global _GEMINI_NEXT_REQUEST_AT
    with _GEMINI_RATE_LOCK:
        now = time.monotonic()
        wait_seconds = max(0.0, _GEMINI_NEXT_REQUEST_AT - now)
        _GEMINI_NEXT_REQUEST_AT = max(now, _GEMINI_NEXT_REQUEST_AT) + GEMINI_MIN_REQUEST_INTERVAL_SECONDS
    if wait_seconds:
        time.sleep(wait_seconds)

def gemini_retry_delay(response: Any, attempt: int) -> float:
    """Honor a server retry hint when present, otherwise use bounded backoff."""
    retry_after = normalize_text(response.headers.get("Retry-After", ""))
    try: return min(60.0, max(1.0, float(retry_after)))
    except ValueError: pass
    try:
        message = normalize_text(response.json().get("error", {}).get("message", ""))
        match = re.search(r"retry(?:Delay| after)?[^\d]*(\d+(?:\.\d+)?)s", message, re.I)
        if match: return min(60.0, max(1.0, float(match.group(1))))
    except (ValueError, AttributeError): pass
    return min(30.0, float(2 ** (attempt + 1)))

def gemini_metadata(api_key: str, row: Dict[str, Any], ai_overview: str) -> Dict[str, Any]:
    """Strict JSON fact extraction from Maps/AI Overview, with no website crawl."""
    fallback = {"owner": "N/A", "operation_time_and_days": "N/A", "doctor_count": None, "branch_count": None, "is_solo": None, "is_collective": None, "nonprofit": None, "private_practice": None, "target_service": None, "red_flags": [], "disallowed_provider_title": None, "outdated_or_insufficient": None, "over_25_years": None, "has_board": None, "status": "not_called"}
    if not api_key.strip(): return fallback
    maps_text = normalize_text(row.get("maps raw text", ""))
    ai_text = normalize_text(ai_overview or "Not shown")
    evidence = f"Google Maps place text:\n{maps_text}\n\nVisible Google AI Overview:\n{ai_text}"
    cache_key = f"{row.get('Practice name', '')}|{row.get('location', '')}|{evidence}"
    with _GEMINI_CACHE_LOCK:
        cached = _GEMINI_CACHE.get(cache_key)
    if cached is not None: return cached
    prompt = """You extract verifiable facts for a US therapy/counseling clinic lead. Use ONLY the supplied Google Maps text and visible Google AI Mode response. Do not use website content, do not browse, and do not invent. Return one JSON object only with exactly these keys: owner (an explicit full personal founder/owner/CEO/clinical-director name, otherwise 'N/A'; NEVER return a role), operation_time_and_days (a concise schedule with days and valid AM/PM times only when explicitly stated, otherwise 'N/A'), doctor_count (integer or null), branch_count (integer or null), is_solo (true/false/null), is_collective (true/false/null), nonprofit (true/false/null), private_practice (true/false/null), target_service (true/false/null only for individual/couples/family/teen therapy, anxiety, depression, trauma, ADHD, bipolar, OCD, DBT, CBT, EMDR, play, art, IFS or listed licenses), red_flags (array containing only actual, directly offered clinic services among intensive outpatient, substance abuse, addiction treatment, medical treatment, peer support, medication management, case management, psychiatric hospital), disallowed_provider_title (true only for an explicit MD, DO, or PMHNP provider; otherwise false/null), outdated_or_insufficient (true only when the evidence explicitly says permanently closed, website down/unavailable/outdated, otherwise false/null), over_25_years (true only for an explicit 25+ years of experience/serving claim, otherwise false/null), has_board (true only for an explicit board of directors, otherwise false/null). For red_flags: include an item ONLY when this clinic directly provides it as a real program/service. Never include it when the evidence says it is not offered, is only a referral to another provider, is offered by a parent/partner/sister organization rather than this clinic, or merely mentions a client condition, a search question, a support group, academic support, or ordinary psychotherapy. In particular, do not treat behavior issues (for example pornography/gaming struggles) as substance abuse or addiction treatment unless a dedicated substance-use/addiction-treatment program is explicitly offered. Do not treat group/family therapy as peer support, or treatment planning as case management. For is_collective: return true ONLY when evidence explicitly calls it a therapist/independent-therapist collective, or says clinicians operate independently under a collective umbrella. Return false for an ordinary group practice, a clinic with a team, a multi-therapist private practice, a collaborative staff, or co-owned practice; those are valid prospects and are not therapist collectives. For nonprofit: return false when the evidence explicitly says it is private, for-profit, or 'not a nonprofit/government agency'. State licensing, Medicaid/public insurance, court approval, government regulation, or a .gov citation do NOT make a private clinic government-owned. Return true only for an affirmative nonprofit, state-owned, government-owned, government-run, or government-funded claim.\n\nEVIDENCE:\n""" + evidence
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={api_key.strip()}"
    try:
        response = None
        for attempt in range(GEMINI_MAX_RETRIES):
            wait_for_gemini_slot()
            response = requests.post(url, json={"contents": [{"parts": [{"text": prompt}]}], "generationConfig": {"responseMimeType": "application/json"}}, timeout=45)
            if response.status_code != 429: break
            if attempt < GEMINI_MAX_RETRIES - 1: time.sleep(gemini_retry_delay(response, attempt))
        if response is None or response.status_code != 200:
            fallback["status"] = f"HTTP {response.status_code if response is not None else 'no response'}"
            return fallback
        raw = response.json().get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.I)
        parsed = json.loads(raw)
        if not isinstance(parsed, dict):
            fallback["status"] = "invalid JSON shape"
            return fallback
        result = {**fallback, **{key: parsed.get(key, fallback[key]) for key in fallback if key != "status"}, "status": "ok"}
        result["owner"] = clean_owner_name(str(result["owner"])) if result["owner"] else "N/A"
        result["operation_time_and_days"] = normalize_text(result["operation_time_and_days"]) or "N/A"
        if result["operation_time_and_days"] != "N/A" and extract_operation_time_from_text(result["operation_time_and_days"]) == "N/A": result["operation_time_and_days"] = "N/A"
        result["red_flags"] = [str(x).lower() for x in result["red_flags"] if str(x).lower() in RED_FLAG_TERMS]
        with _GEMINI_CACHE_LOCK:
            _GEMINI_CACHE[cache_key] = result
        return result
    except (requests.RequestException, ValueError, KeyError, IndexError) as exc:
        fallback["status"] = f"error: {type(exc).__name__}"
        return fallback

def merge_gemini_evidence(base: Evidence, metadata: Dict[str, Any]) -> Evidence:
    if metadata.get("status") != "ok": return base
    if metadata.get("owner") and metadata["owner"] != "N/A": base.owner = metadata["owner"]
    if isinstance(metadata.get("doctor_count"), int): base.doctor_count = metadata["doctor_count"]
    if isinstance(metadata.get("branch_count"), int): base.branch_count = metadata["branch_count"]
    for attr, key in (("is_solo", "is_solo"), ("is_collective", "is_collective"), ("nonprofit", "nonprofit")):
        if metadata.get(key) in (True, False): setattr(base, attr, metadata[key])
    for attr, key in (("old_or_insufficient", "outdated_or_insufficient"), ("over_25_years", "over_25_years"), ("has_board", "has_board"), ("disallowed_provider_title", "disallowed_provider_title")):
        if metadata.get(key) in (True, False): setattr(base, attr, metadata[key])
    for attr, key in (("private_practice", "private_practice"), ("target_service", "target_service")):
        if metadata.get(key) in (True, False): setattr(base, attr, metadata[key])
    base.red_flags = list(metadata.get("red_flags", []))
    return base

def format_export(rows: List[Dict[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame([{ "Practice Name": r.get("Practice name", "N/A"), "Website Link": r.get("website link", "N/A"), "Phone Number": r.get("phone number", "N/A"), "Location (City Ne Only)": r.get("location", "N/A"), "Operation Time and Days": r.get("operation time and days", "N/A"), "Rating Star": r.get("rating star", "N/A"), "Owner": r.get("Owner's name", "N/A")} for r in rows], columns=EXPORT_COLUMNS)

def lead_key(name: Any, city: Any) -> Tuple[str, str]:
    return normalize_text(name).lower(), normalize_text(city).lower()

def candidate_id(sheet: str, row: Dict[str, Any]) -> str:
    name, city = lead_key(row.get("Practice name", ""), row.get("location", ""))
    return f"{sheet}\x1f{name}\x1f{city}"

def existing_lead_keys(workbook_bytes: bytes) -> set:
    found = set()
    with pd.ExcelFile(io.BytesIO(workbook_bytes)) as book:
        for sheet in book.sheet_names:
            df = pd.read_excel(book, sheet_name=sheet, dtype=str).fillna(""); cols = {str(c).lower().strip(): c for c in df.columns}
            if "practice name" in cols:
                fallback_city = sheet.split(",", 1)[0].strip()
                city_column = cols.get("location (city ne only)") or cols.get("location")
                for _, row in df.iterrows(): found.add(lead_key(row[cols["practice name"]], row[city_column] if city_column else fallback_city))
    return found

def discover_city_sheets(workbook_bytes: bytes, fallback_state: str = "") -> List[Dict[str, str]]:
    """Find lead sheets and derive city/state from names such as `Provo, UT`.

    Only sheets carrying the same seven lead headers are scheduled. This keeps
    cover sheets, notes and unrelated tabs untouched in the downloaded file.
    """
    city_sheets: List[Dict[str, str]] = []
    with pd.ExcelFile(io.BytesIO(workbook_bytes)) as book:
        for sheet_name in book.sheet_names:
            headers = [normalize_text(x).lower() for x in pd.read_excel(book, sheet_name=sheet_name, nrows=0).columns]
            if "practice name" not in headers or "phone number" not in headers:
                continue
            city, separator, state = sheet_name.partition(",")
            city = city.strip()
            state = state.strip() if separator else fallback_state.strip()
            if city:
                city_sheets.append({"sheet_name": sheet_name, "city": city, "state": state})
    return city_sheets

def unique_excel_sheet_name(preferred: str, used_names: Iterable[str]) -> str:
    """Create a valid, collision-free Excel tab name (Excel allows 31 chars)."""
    base = re.sub(r"[\\\\/*?:\[\]]", "-", normalize_text(preferred)) or "Leads"
    base = base[:31]
    used = {name.casefold() for name in used_names}
    candidate, number = base, 2
    while candidate.casefold() in used:
        suffix = f" ({number})"
        candidate = base[:31 - len(suffix)] + suffix
        number += 1
    return candidate

def append_rows_preserving_template(template: bytes, rows_by_sheet: Dict[str, pd.DataFrame]) -> bytes:
    workbook = load_workbook(io.BytesIO(template))
    for wanted, rows in rows_by_sheet.items():
        if rows.empty: continue
        ws = workbook[wanted] if wanted in workbook.sheetnames else workbook.create_sheet(wanted)
        # Preserve every existing row and column. Sheet discovery is deliberately
        # lenient, so an older workbook may have only some export columns or use
        # different capitalization. Add only missing columns instead of replacing
        # the sheet, which would destroy the user's existing leads.
        existing_locations: Dict[str, int] = {}
        for column in range(1, ws.max_column + 1):
            header = normalize_text(ws.cell(1, column).value)
            if header:
                existing_locations.setdefault(header.casefold(), column)
        next_column = 1 if not existing_locations and ws.cell(1, 1).value is None else ws.max_column + 1
        locations: Dict[str, int] = {}
        for header in EXPORT_COLUMNS:
            column = existing_locations.get(header.casefold())
            if column is None:
                column = next_column
                next_column += 1
                ws.cell(1, column, header).font = Font(name="Arial", bold=True)
                existing_locations[header.casefold()] = column
            locations[header] = column
        added_row_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        for record in rows.to_dict("records"):
            n = ws.max_row + 1
            for header in EXPORT_COLUMNS:
                cell = ws.cell(n, locations[header], record.get(header, "N/A"))
                cell.fill = added_row_fill
    output = io.BytesIO(); workbook.save(output); return output.getvalue()

def new_workbook(rows_by_sheet: Dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, rows in rows_by_sheet.items(): rows.to_excel(writer, sheet_name=sheet, index=False)
    return output.getvalue()

def run_job(driver: webdriver.Chrome, city: str, state: str, keyword: str, limit: int, known: set, progress: Any, gemini_api_key: str = "", captcha_wait_seconds: int = 300, on_accept: Optional[Any] = None, should_stop: Optional[Any] = None, known_lock: Optional[Any] = None) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    accepted: List[Dict[str, Any]] = []; debug: List[Dict[str, Any]] = []; places = maps_search_urls(driver, city, state, keyword, limit, progress, captcha_wait_seconds)
    for position, (url, name_hint) in enumerate(places, 1):
        if should_stop and should_stop(): break
        progress(f"{city}: {keyword} — {position}/{len(places)}")
        row: Optional[Dict[str, Any]] = None
        try:
            row = extract_maps_place_with_retry(driver, url, city, name_hint, progress, captcha_wait_seconds); key = lead_key(row["Practice name"], row["location"])
            if row["Practice name"] == "N/A":
                debug.append({"Practice": "N/A", "Keep": False, "Filter result": "SKIP: không đọc được tên từ Google Maps", "Owner": "N/A", "Operation Time": "N/A", "Therapists": "UNKNOWN", "Private practice": "UNKNOWN", "Target services": "UNKNOWN", "AI Mode evidence": "N/A", "Gemini": "not called", "Reasons": "Google Maps listing could not be read"})
                continue
            if known_lock:
                with known_lock:
                    if key in known: continue
                    known.add(key)
            elif key in known:
                # Duplicate leads are deliberately skipped without cluttering Debug.
                continue
            ai_mode_evidence = google_ai_overview(driver, row["Practice name"], city, state, progress, captcha_wait_seconds)
            if not ai_mode_evidence:
                # Never ask Gemini to decide from Maps alone when the required AI
                # Mode evidence is missing or still streaming. Let another keyword
                # encounter retry this clinic instead of permanently deduplicating it.
                if known_lock:
                    with known_lock:
                        known.discard(key)
                else:
                    known.discard(key)
                debug.append({"Practice": row["Practice name"], "Keep": False, "Filter result": "RETRY: AI Mode chưa hoàn tất", "Owner": "N/A", "Operation Time": row["operation time and days"], "Therapists": "UNKNOWN", "Private practice": "UNKNOWN", "Target services": "UNKNOWN", "AI Mode evidence": "", "Gemini": "not called", "Reasons": "AI Mode did not produce a complete stable answer after retries"})
                continue
            overview = f"AI MODE (structured clinic screening query):\n{ai_mode_evidence}"
            metadata = gemini_metadata(gemini_api_key, row, overview); e = merge_gemini_evidence(Evidence(), metadata); v = evidence_as_verification(e); row["Owner's name"] = e.owner
            gemini_hours = metadata.get("operation_time_and_days", "N/A")
            if metadata.get("status") == "ok" and gemini_hours != "N/A": row["operation time and days"] = gemini_hours
            keep = should_keep_in_final_output(v)
            debug.append({"Practice": row["Practice name"], "Keep": keep, "Filter result": filter_result(v), "Owner": e.owner, "Operation Time": row["operation time and days"], "Therapists": str(e.doctor_count) if e.doctor_count is not None else "UNKNOWN", "Private practice": "YES" if e.private_practice is True else ("NO" if e.private_practice is False else "UNKNOWN"), "Target services": "YES" if e.target_service is True else ("NO" if e.target_service is False else "UNKNOWN"), "AI Mode evidence": ai_mode_evidence, "Gemini": metadata.get("status"), "Reasons": "; ".join(e.red_flags) or "eligible / insufficient evidence", "_export_row": dict(row)})
            if not known_lock: known.add(key)
            if keep:
                accepted.append(row)
                if on_accept: on_accept(row)
        except (TimeoutException, WebDriverException) as exc:
            practice = row.get("Practice name", "N/A") if row else "N/A"
            debug.append({"Practice": practice, "Keep": False, "Filter result": f"Google UI error after retry: {type(exc).__name__}", "Owner": "N/A", "Operation Time": row.get("operation time and days", "N/A") if row else "N/A", "Therapists": "UNKNOWN", "Private practice": "UNKNOWN", "Target services": "UNKNOWN", "AI Mode evidence": "N/A", "Gemini": "not called", "Reasons": f"Google UI error: {type(exc).__name__}"})
    return accepted, debug

def checkpoint_output(source: bytes, rows_by_sheet: Dict[str, List[Dict[str, Any]]]) -> bytes:
    exports = {sheet: format_export(rows) for sheet, rows in rows_by_sheet.items()}
    return append_rows_preserving_template(source, exports) if source else new_workbook(exports)

def save_checkpoint_locked(job: Dict[str, Any]) -> None:
    output = checkpoint_output(job["source"], job["rows_by_sheet"])
    Path(job["checkpoint_path"]).write_bytes(output)
    job["checkpoint_bytes"] = output

def checkpoint_directory() -> Path:
    """Return a writable persistent checkpoint location for this platform."""
    if sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support" / "ClinicLeadCollector" / "checkpoints"
    else:
        base = Path(os.environ.get("LOCALAPPDATA", Path.home())) / "ClinicLeadCollector" / "checkpoints"
    base.mkdir(parents=True, exist_ok=True)
    checkpoints = sorted(base.glob("clinic_leads_checkpoint_*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    for old_file in checkpoints[MAX_SAVED_CHECKPOINTS:]:
        try: old_file.unlink()
        except OSError: pass
    return base

def installed_app_directory() -> Path:
    if not getattr(sys, "frozen", False):
        return Path(__file__).resolve().parent
    executable = Path(sys.executable).resolve()
    if sys.platform == "darwin":
        for parent in executable.parents:
            if parent.suffix.casefold() == ".app":
                return parent
    return executable.parent

def installed_app_version_path() -> Path:
    app_directory = installed_app_directory()
    if sys.platform == "darwin" and app_directory.suffix.casefold() == ".app":
        return app_directory / "Contents" / "Resources" / "version.txt"
    return app_directory / "version.txt"

def installed_app_version() -> str:
    try: return installed_app_version_path().read_text(encoding="utf-8-sig").strip() or "dev"
    except OSError: return "dev"

def updater_executable() -> Path:
    app_directory = installed_app_directory()
    if sys.platform == "darwin":
        return app_directory / "Contents" / "Resources" / "Clinic Lead Updater"
    return app_directory / "Clinic Lead Updater.exe"

def promote_updater_payload() -> None:
    """Install the updater shipped as a sidecar while the updater is not running.

    Older updater builds deliberately skip their own executable. They do copy
    this differently named payload, allowing the newly restarted main app to
    promote it and keeping the updater itself upgradeable.
    """
    if not getattr(sys, "frozen", False) or sys.platform != "win32": return
    app_directory = installed_app_directory()
    payload = app_directory / "Clinic Lead Updater Payload.exe"
    updater = app_directory / "Clinic Lead Updater.exe"
    if not payload.exists(): return
    # The previous updater launches this app just before its own process exits,
    # so Windows may keep the old executable locked for a brief moment.
    for _ in range(20):
        try:
            os.replace(payload, updater)
            return
        except OSError:
            time.sleep(0.1)
    # A failed promotion must not prevent the collector from starting. The
    # payload stays in place so the next launch can retry.

def available_release() -> Optional[Dict[str, str]]:
    """Read the public GitHub Release metadata; never send user/API data."""
    asset_name = UPDATE_ASSET_NAMES.get(sys.platform)
    if not asset_name: return None
    try:
        response = requests.get(f"https://api.github.com/repos/{UPDATE_REPOSITORY}/releases/latest", timeout=6)
        if response.status_code != 200: return None
        release = response.json(); tag = normalize_text(release.get("tag_name", ""))
        for asset in release.get("assets", []):
            if asset.get("name") == asset_name:
                return {"version": tag, "url": normalize_text(asset.get("browser_download_url", ""))}
    except (requests.RequestException, ValueError, AttributeError):
        pass
    return None

def render_update_control() -> None:
    """Offer a user-confirmed desktop update when a newer public release exists."""
    if not getattr(sys, "frozen", False): return
    current = installed_app_version(); release = available_release()
    if not release or not release["url"] or release["version"] == current: return
    st.sidebar.info(f"Có bản cập nhật mới: {release['version']} (đang dùng {current})")
    if st.sidebar.button("Cập nhật app và khởi động lại", type="primary"):
        updater = updater_executable()
        if not updater.exists():
            st.sidebar.error("Không tìm thấy updater trong thư mục app.")
            return
        subprocess.Popen([str(updater), "--target", str(installed_app_directory()), "--url", release["url"], "--parent-pid", str(os.getpid())], close_fds=True)
        os._exit(0)

def start_background_job(config: Dict[str, Any]) -> Dict[str, Any]:
    """Run a small visible-Chrome worker pool and checkpoint every accepted lead."""
    global _GEMINI_NEXT_REQUEST_AT
    with _GEMINI_CACHE_LOCK:
        _GEMINI_CACHE.clear()
    with _GEMINI_RATE_LOCK:
        _GEMINI_NEXT_REQUEST_AT = 0.0
    checkpoint_dir = checkpoint_directory()
    job: Dict[str, Any] = {
        **config, "lock": threading.Lock(), "stop_event": threading.Event(), "running": True,
        "message": "Đang chuẩn bị…", "error": "",
        "known_lock": threading.Lock(), "captcha_workers": {},
        "rows_by_sheet": {sheet: [] for _, _, sheet in config["jobs"]}, "debug": [], "candidates": {},
        "checkpoint_path": str(checkpoint_dir / f"clinic_leads_checkpoint_{uuid.uuid4().hex[:8]}.xlsx"),
        "checkpoint_bytes": b"", "captcha_active": False, "captcha_notified": False, "captcha_sound_played": False,
    }
    with job["lock"]:
        save_checkpoint_locked(job)
    def worker() -> None:
        tasks: Queue = Queue()
        for city, state, sheet in job["jobs"]:
            for keyword in job["keywords"]:
                tasks.put((city, state, sheet, keyword))
        worker_count = min(max(1, int(job.get("parallel_workers", 1))), 5, tasks.qsize() or 1)

        def browser_worker(worker_number: int) -> None:
            driver: Optional[webdriver.Chrome] = None
            try:
                driver = build_driver(job["headless"])
                while not job["stop_event"].is_set():
                    try: city, state, sheet, keyword = tasks.get_nowait()
                    except Empty: break
                    try:
                        def progress(message: str) -> None:
                            with job["lock"]:
                                is_captcha = "CAPTCHA" in message.upper()
                                job["captcha_workers"][worker_number] = is_captcha
                                job["captcha_active"] = any(job["captcha_workers"].values())
                                if not job["captcha_active"]:
                                    job["captcha_notified"] = False; job["captcha_sound_played"] = False
                                elif job.get("captcha_notifications") and not job["captcha_sound_played"]:
                                    job["captcha_sound_played"] = True
                                    if winsound:
                                        threading.Thread(target=lambda: winsound.MessageBeep(winsound.MB_ICONEXCLAMATION), daemon=True).start()
                                job["message"] = f"Chrome {worker_number}/{worker_count}: {message}"
                        def on_accept(row: Dict[str, Any], target_sheet: str = sheet) -> None:
                            with job["lock"]:
                                job["rows_by_sheet"][target_sheet].append(row)
                                save_checkpoint_locked(job)
                                job["message"] = f"Đã lưu checkpoint: {sum(len(x) for x in job['rows_by_sheet'].values())} lead mới."
                        _, details = run_job(driver, city, state, keyword, job["limit"], job["known"], progress, job["gemini_api_key"], job["captcha_wait_seconds"], on_accept, job["stop_event"].is_set, job["known_lock"])
                        with job["lock"]:
                            for item in details:
                                export_row = item.pop("_export_row", None)
                                if export_row:
                                    item["Candidate ID"] = candidate_id(sheet, export_row)
                                    job["candidates"][item["Candidate ID"]] = export_row
                                job["debug"].append({**item, "Sheet": sheet, "City": city, "State": state})
                    except Exception as exc:
                        with job["lock"]:
                            job["debug"].append({"Practice": "N/A", "Keep": False, "Filter result": f"Chrome worker error: {type(exc).__name__}", "Reasons": str(exc), "Sheet": sheet, "City": city, "State": state})
                    finally:
                        tasks.task_done()
            except Exception as exc:
                with job["lock"]:
                    job["error"] = f"Chrome {worker_number}: {type(exc).__name__}: {exc}"
            finally:
                if driver:
                    try: driver.quit()
                    except Exception: pass
                with job["lock"]:
                    job["captcha_workers"].pop(worker_number, None)
                    job["captcha_active"] = any(job["captcha_workers"].values())

        workers = [threading.Thread(target=browser_worker, args=(number,), daemon=True, name=f"clinic-browser-{number}") for number in range(1, worker_count + 1)]
        for browser_thread in workers: browser_thread.start()
        for browser_thread in workers: browser_thread.join()
        try:
            with job["lock"]:
                save_checkpoint_locked(job)
                job["message"] = "Đã dừng và lưu checkpoint." if job["stop_event"].is_set() else "Hoàn tất và đã lưu file cuối."
        except Exception as exc:
            with job["lock"]:
                save_checkpoint_locked(job)
                job["error"] = f"{type(exc).__name__}: {exc}"
                job["message"] = "Có lỗi, nhưng checkpoint đã được lưu."
        finally:
            with job["lock"]:
                job["running"] = False
    threading.Thread(target=worker, daemon=True, name="clinic-scraper").start()
    return job

def apply_debug_keep_selection(job: Dict[str, Any], edited_debug: pd.DataFrame) -> int:
    """Make the editable Debug Keep checkboxes the source of output selection."""
    changed = 0
    with job["lock"]:
        for _, debug_row in edited_debug.iterrows():
            selected = bool(debug_row.get("Keep", False)); item_id = normalize_text(debug_row.get("Candidate ID", ""))
            candidate = job["candidates"].get(item_id)
            if not candidate: continue
            sheet = normalize_text(debug_row.get("Sheet", ""))
            if sheet not in job["rows_by_sheet"]: continue
            for stored_debug in job["debug"]:
                if stored_debug.get("Candidate ID") == item_id:
                    stored_debug["Keep"] = selected
                    break
            rows = job["rows_by_sheet"][sheet]
            present = any(candidate_id(sheet, row) == item_id for row in rows)
            if selected and not present:
                rows.append(dict(candidate)); changed += 1
            elif not selected and present:
                job["rows_by_sheet"][sheet] = [row for row in rows if candidate_id(sheet, row) != item_id]
                changed += 1
        if changed:
            save_checkpoint_locked(job)
            job["message"] = f"Đã cập nhật lựa chọn Debug và lưu checkpoint ({sum(len(rows) for rows in job['rows_by_sheet'].values())} lead mới)."
    return changed

def _render_background_job(job: Dict[str, Any]) -> None:
    with job["lock"]:
        message, running, error = job["message"], job["running"], job["error"]
        stop_requested = job["stop_event"].is_set()
        bytes_now = job["checkpoint_bytes"]
        summary = pd.DataFrame([{"Sheet": sheet, "Lead mới đã lưu": len(rows)} for sheet, rows in job["rows_by_sheet"].items()])
        debug = list(job["debug"]); notify = bool(job.get("captcha_notifications")) and job["captcha_active"] and not job["captcha_notified"]
        if notify: job["captcha_notified"] = True
    st.info(message)
    if notify:
        st.warning("CAPTCHA đang chặn Google. Hãy mở cửa sổ Chrome và xác minh thủ công để bot tiếp tục.")
        components.html("""<script>
          try { if (Notification.permission === 'default') Notification.requestPermission();
            if (Notification.permission === 'granted') new Notification('Clinic scraper', {body: 'Google đang yêu cầu CAPTCHA.'});
            const c = new (window.AudioContext || window.webkitAudioContext)(); const o = c.createOscillator(); const g = c.createGain();
            o.connect(g); g.connect(c.destination); o.frequency.value = 880; g.gain.value = 0.08; o.start(); setTimeout(() => { o.stop(); c.close(); }, 450);
          } catch (_) {} </script>""", height=0)
    if error: st.error(error)
    if running and st.button("Dừng và lưu checkpoint ngay", type="secondary"):
        job["stop_event"].set(); stop_requested = True; st.warning("Đã gửi yêu cầu dừng. App sẽ lưu xong lead đang xử lý rồi dừng.")
    if running and stop_requested:
        if st.button("Quay lại màn hình chính ngay", type="secondary", key="return_while_stopping"):
            # Keep the worker alive so its checkpoint is still finalized; only
            # detach this status panel from the Streamlit session.
            st.session_state["stopping_scrape_job"] = job
            st.session_state.pop("scrape_job", None)
            st.rerun()
    st.subheader("📥 File Excel đã lưu")
    st.download_button("TẢI FILE EXCEL NGAY", data=bytes_now, file_name="clinic_leads_checkpoint.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
    st.dataframe(summary, width="stretch")
    with st.expander("Evidence debug (AI Mode + Gemini)"):
        debug_frame = pd.DataFrame(debug)
        if debug_frame.empty:
            st.caption("Chưa có lead nào được kiểm tra.")
        else:
            editable_columns = [column for column in debug_frame.columns if column != "Keep"]
            edited_debug = st.data_editor(
                debug_frame, hide_index=True, width="stretch", key="evidence_debug_editor",
                disabled=editable_columns,
                column_config={"Keep": st.column_config.CheckboxColumn("Keep", help="Tick để lưu lead này vào file kết quả."), "Candidate ID": None},
            )
            if apply_debug_keep_selection(job, edited_debug):
                st.success("Đã lưu lựa chọn vào checkpoint. Các dòng mới trong file Excel upload sẽ được tô vàng.")
    if running:
        st.caption("Trạng thái và bảng debug tự cập nhật mỗi 3 giây; bạn vẫn có thể tải checkpoint hoặc dừng bất cứ lúc nào.")
    else:
        st.success("Lượt chạy đã kết thúc. File checkpoint phía trên là file kết quả.")
        st.divider()
        st.caption("Bắt đầu một lượt hoàn toàn mới")
        _, restart_col, _ = st.columns([1, 2, 1])
        with restart_col:
            if st.button("Tạo lượt chạy mới", type="primary", key="restart_finished"):
                st.session_state.pop("scrape_job", None); st.rerun()

def show_background_job(job: Dict[str, Any]) -> None:
    """Auto-refresh the running panel when the installed Streamlit supports fragments."""
    if hasattr(st, "fragment"):
        @st.fragment(run_every=3)
        def live_job_panel() -> None:
            _render_background_job(job)
        live_job_panel()
    else:
        _render_background_job(job)

def main() -> None:
    st.title("🏥 Tìm kiếm phòng khám")
    st.caption("Tìm trên Google Maps, kiểm tra bằng AI Mode và lưu kết quả vào Excel.")
    promote_updater_payload()
    render_update_control()
    active_job = st.session_state.get("scrape_job")
    if active_job:
        show_background_job(active_job)
        return
    stopping_job = st.session_state.get("stopping_scrape_job")
    if stopping_job:
        with stopping_job["lock"]:
            still_stopping = stopping_job["running"]
        if still_stopping:
            st.info("Lượt trước đang dừng và lưu kết quả tạm thời. Bạn vẫn có thể chuẩn bị lượt tiếp theo.")
        else:
            st.session_state.pop("stopping_scrape_job", None)
    checkpoint_dir = checkpoint_directory()
    old_checkpoints = sorted(checkpoint_dir.glob("clinic_leads_checkpoint_*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True) if checkpoint_dir.exists() else []
    if old_checkpoints:
        latest = old_checkpoints[0]
        st.download_button("Khôi phục kết quả tạm từ lượt trước", data=latest.read_bytes(), file_name="clinic_leads_recovered.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", help="Dùng khi lượt chạy trước bị tắt giữa chừng.")

    st.subheader("1. Dữ liệu cũ")
    upload = st.file_uploader(
        "Tải file Excel đang dùng",
        type=["xlsx"],
        help="App giữ nguyên dữ liệu cũ, tránh lặp lại lead đã có và thêm kết quả mới vào file. Không có file cũ thì bỏ qua bước này.",
    )

    st.subheader("2. Khu vực cần tìm")
    city_col, state_col = st.columns([2, 1])
    with city_col:
        city = st.text_input(
            "Thành phố",
            "",
            placeholder="Ví dụ: Provo, Orem, Sandy",
            help="Có thể nhập nhiều thành phố, ngăn cách bằng dấu phẩy.",
        )
    with state_col:
        state = st.text_input(
            "Bang",
            "",
            placeholder="Ví dụ: UT",
            help="Bang dùng cho các thành phố vừa nhập và sheet chỉ có tên thành phố. Ví dụ: UT = Utah, CA = California. Nếu sheet đã có dạng 'Provo, UT' thì app tự nhận bang.",
        )

    st.subheader("3. Kết nối Gemini")
    gemini_api_key = st.text_input(
        "Gemini API key",
        type="password",
        placeholder="Dán API key vào đây",
        help="Bắt buộc để AI kiểm tra và chuẩn hoá thông tin. Key chỉ được dùng trong lượt chạy này.",
    )

    with st.expander("⚙️ Cài đặt nâng cao", expanded=False):
        st.caption("Các giá trị mặc định đã phù hợp cho hầu hết trường hợp.")
        keywords = st.text_area("Từ khoá tìm kiếm (mỗi dòng một từ khoá)", "counseling center\nmental health clinic\ntherapy practice")
        limit = st.number_input("Số kết quả tối đa cho mỗi từ khoá", 1, 200, 100)
        captcha_wait_seconds = st.number_input("Thời gian chờ xác minh CAPTCHA (giây)", min_value=30, max_value=900, value=300, step=30, help="Khi Google yêu cầu xác minh, app sẽ chờ bạn giải CAPTCHA trong khoảng thời gian này.")
        captcha_notifications = st.checkbox("Thông báo khi cần xác minh CAPTCHA", value=True)
    source = upload.getvalue() if upload else b""
    detected = discover_city_sheets(source, state) if source else []
    selected_sheets: List[str] = []
    if detected:
        sheet_to_label = {item["sheet_name"]: f"{item['sheet_name']}  →  {item['city']}, {item['state'] or 'state chưa rõ'}" for item in detected}
        selected_sheets = st.multiselect("Các sheet sẽ tiếp tục tìm", options=list(sheet_to_label), default=list(sheet_to_label), format_func=lambda name: sheet_to_label[name])
        st.success(f"Đã nhận {len(detected)} sheet có dữ liệu lead. Dữ liệu cũ và các sheet khác sẽ được giữ nguyên.")
    elif upload:
        st.warning("File chưa có sheet lead hợp lệ (cần cột Practice Name và Phone Number). Bạn vẫn có thể tìm theo thành phố đã nhập.")
    run_manual = bool(city.strip())
    if detected and city.strip():
        st.info("App sẽ tìm cả các sheet đã chọn và các thành phố bạn vừa nhập.")
    st.subheader("4. Bắt đầu")
    if not st.button("🔍 BẮT ĐẦU TÌM KIẾM", type="primary", use_container_width=True): return
    if not keywords.strip(): st.error("Nhập ít nhất một keyword."); return
    if not gemini_api_key.strip(): st.error("Vui lòng nhập Gemini API key trước khi chạy."); return
    jobs: List[Tuple[str, str, str]] = [(item["city"], item["state"], item["sheet_name"]) for item in detected if item["sheet_name"] in selected_sheets]
    used_sheet_names = [item[2] for item in jobs]
    if source:
        with pd.ExcelFile(io.BytesIO(source)) as book:
            used_sheet_names.extend(book.sheet_names)
    if run_manual and city.strip():
        seen_manual_cities = set()
        for manual_city in (item.strip() for item in city.split(",") if item.strip()):
            city_key = manual_city.lower()
            if city_key in seen_manual_cities: continue
            seen_manual_cities.add(city_key)
            preferred_sheet = f"{manual_city}, {state.strip()}" if state.strip() else manual_city
            if all(sheet.casefold() != preferred_sheet.casefold() for _, _, sheet in jobs):
                manual_sheet = unique_excel_sheet_name(preferred_sheet, used_sheet_names)
                jobs.append((manual_city, state.strip(), manual_sheet)); used_sheet_names.append(manual_sheet)
    if not jobs: st.error("Vui lòng chọn ít nhất một sheet hoặc nhập thành phố cần tìm."); return
    missing_state_cities = [job_city for job_city, job_state, _ in jobs if not job_state]
    if missing_state_cities:
        preview = ", ".join(missing_state_cities[:3])
        st.error(f"Vui lòng nhập mã bang cho: {preview}. Nếu file có nhiều bang, hãy đặt tên sheet theo dạng 'Thành phố, Bang', ví dụ 'Provo, UT'.")
        return
    st.session_state["scrape_job"] = start_background_job({
        "source": source, "jobs": jobs, "keywords": [x.strip() for x in keywords.splitlines() if x.strip()],
        "limit": int(limit), "parallel_workers": 1, "headless": False, "known": existing_lead_keys(source) if source else set(),
        "gemini_api_key": gemini_api_key, "captcha_wait_seconds": int(captcha_wait_seconds), "captcha_notifications": captcha_notifications,
    })
    st.rerun()

if __name__ == "__main__": main()
